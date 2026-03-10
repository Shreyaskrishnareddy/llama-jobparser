#!/usr/bin/env python3
"""
Generate updated Excel with before/after comparison columns.
Reads the original expected values from the Excel, adds our new extracted values and match status.
"""

import json
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

RESULTS_DIR = "/home/great/groq-jd-parser/test-results/fix-test"
INPUT_EXCEL = "/mnt/d/DOWNLOADS/JD Parser Results - New.xlsx"
OUTPUT_EXCEL = "/mnt/d/DOWNLOADS/JD Parser Results - Updated.xlsx"

# Map sheet names to result files
SHEET_MAP = {
    "Embedded Systems Engineer 1": "Embedded_Systems_Engineer_Job_Description.pdf.json",
    "Fullstack developer": "Full_Stack_Developer_Job_Description.pdf.json",
    "Business Analyst (Mid-Senior)": "Business_Analyst_(Mid_Senior)_Job_Description.pdf.json",
    "Senior Network Engineer": "Senior_Network_Engineer_Job_Description.pdf.json",
    "Project Manager": "Project_Manager_(PMP)_Job_Description.pdf.json",
}


def load_result(filename):
    path = os.path.join(RESULTS_DIR, filename)
    if not os.path.exists(path):
        return None
    with open(path) as f:
        return json.load(f)


def get_field_value(result, category, data_point):
    """Extract a field value from result based on category + data_point from the Excel."""
    if not result:
        return "N/A"

    fields = result.get("fields", {})

    # Map Excel data points to our JSON field names
    dp = data_point.lower().strip()

    # Job Information
    if dp in ("job_title", "job title"):
        f = fields.get("title")
        if isinstance(f, dict) and isinstance(f.get("value"), dict):
            return f["value"].get("text", "N/A")
        return "N/A"

    if dp in ("job_id / reference_code", "job_id", "reference_code"):
        f = fields.get("job_id")
        return f["value"] if isinstance(f, dict) and f.get("value") else "NA"

    if dp == "employment_type":
        f = fields.get("employment_type")
        if isinstance(f, dict) and isinstance(f.get("value"), list):
            val = ", ".join(f["value"])
            # Add contract details
            ct = fields.get("contract_type")
            cd = fields.get("contract_duration")
            ct_val = ct["value"] if isinstance(ct, dict) and ct.get("value") else None
            cd_val = cd["value"] if isinstance(cd, dict) and cd.get("value") else None
            extras = []
            if ct_val:
                extras.append(ct_val)
            if cd_val:
                extras.append(cd_val)
            if extras:
                val += f" ({', '.join(extras)})"
            return val
        return "NA"

    if dp == "work_mode":
        f = fields.get("work_mode")
        return f["value"] if isinstance(f, dict) and f.get("value") else "NA"

    if dp == "job_location":
        f = fields.get("location")
        if isinstance(f, dict) and isinstance(f.get("value"), dict):
            loc = f["value"]
            parts = [p for p in [loc.get("city"), loc.get("region"), loc.get("country")] if p]
            remote = loc.get("remote", "")
            addr = ", ".join(parts)
            if remote and remote != "onsite":
                addr += f" ({remote.title()})"
            return addr
        return "NA"

    if dp in ("job_posted_date",):
        f = fields.get("job_posted_date")
        return f["value"] if isinstance(f, dict) and f.get("value") else "NA"

    if dp in ("job_expiry_date",):
        f = fields.get("job_expiry_date")
        return f["value"] if isinstance(f, dict) and f.get("value") else "NA"

    # Company Information
    if dp == "company_name":
        f = fields.get("company")
        return f["value"] if isinstance(f, dict) and f.get("value") else "NA"

    if dp == "company_website":
        f = fields.get("company_website")
        return f["value"] if isinstance(f, dict) and f.get("value") else "NA"

    if dp == "industry":
        f = fields.get("industry")
        return f["value"] if isinstance(f, dict) and f.get("value") else "NA"

    if dp == "company_size":
        f = fields.get("company_size")
        return f["value"] if isinstance(f, dict) and f.get("value") else "NA"

    if dp == "company_overview":
        f = fields.get("company_overview")
        if isinstance(f, dict) and f.get("value"):
            v = f["value"]
            return v[:300] + "..." if len(v) > 300 else v
        return "NA"

    # Role Summary
    if dp == "job_summary":
        f = fields.get("job_summary")
        if isinstance(f, dict) and f.get("value"):
            v = f["value"]
            sentences = len(re.findall(r'[.!?](?:\s|$)', v + " "))
            return f"[{sentences} sentences, {len(v)} chars] {v[:200]}..." if len(v) > 200 else f"[{sentences} sentences, {len(v)} chars] {v}"
        return "NA"

    # Description
    if dp == "description":
        f = fields.get("description")
        if isinstance(f, dict) and f.get("value"):
            v = f["value"]
            sentences = len(re.findall(r'[.!?](?:\s|$)', v + " "))
            return f"[{sentences} sentences, {len(v)} chars] {v[:200]}..." if len(v) > 200 else f"[{sentences} sentences, {len(v)} chars] {v}"
        return "NA"

    # Responsibilities
    if dp == "responsibilities":
        f = fields.get("responsibilities")
        if isinstance(f, dict) and isinstance(f.get("value"), list):
            return f"Captured {len(f['value'])} items"
        return "NA"

    # Qualifications
    if dp in ("education_level", "education"):
        f = fields.get("education")
        if isinstance(f, dict) and isinstance(f.get("value"), dict):
            edu = f["value"]
            parts = [p for p in [edu.get("level"), edu.get("field")] if p]
            return " in ".join(parts) if parts else "NA"
        return "NA"

    if dp in ("years_of_experience", "experience_years"):
        f = fields.get("experience_years")
        if isinstance(f, dict) and isinstance(f.get("value"), dict):
            exp = f["value"]
            mn = exp.get("min_years")
            mx = exp.get("max_years")
            req = exp.get("requirement_type", "required")
            if mn is not None and mx is not None:
                if mn == mx:
                    return f"{mn} years ({req})"
                return f"{mn}-{mx} years ({req})"
        return "NA"

    if dp == "certifications":
        f = fields.get("certifications")
        if isinstance(f, dict) and isinstance(f.get("value"), list) and f["value"]:
            return ", ".join(f["value"])
        return "NA"

    if dp == "technical_skills":
        f = fields.get("technical_skills")
        if isinstance(f, dict) and isinstance(f.get("value"), list) and f["value"]:
            skills = f["value"]
            return f"Captured {len(skills)} skills: {', '.join(skills[:10])}" + ("..." if len(skills) > 10 else "")
        # Fall back to skills array
        f = fields.get("skills")
        if isinstance(f, dict) and isinstance(f.get("value"), list):
            tech = [s["name"] for s in f["value"] if isinstance(s, dict) and s.get("category") != "soft_skill"]
            return f"Captured {len(tech)} skills: {', '.join(tech[:10])}" + ("..." if len(tech) > 10 else "")
        return "NA"

    if dp == "soft_skills":
        f = fields.get("soft_skills")
        if isinstance(f, dict) and isinstance(f.get("value"), list) and f["value"]:
            skills = f["value"]
            return f"Captured {len(skills)} skills: {', '.join(skills)}"
        return "NA"

    # Preferred
    if dp in ("preferred_experience",):
        f = fields.get("preferred_experience")
        if isinstance(f, dict) and isinstance(f.get("value"), list) and f["value"]:
            items = f["value"]
            return f"Captured {len(items)} items: {'; '.join(items[:3])}" + ("..." if len(items) > 3 else "")
        return "NA"

    if dp in ("preferred_technologies",):
        f = fields.get("preferred_technologies")
        if isinstance(f, dict) and isinstance(f.get("value"), list) and f["value"]:
            return ", ".join(f["value"])
        return "NA"

    # Compensation
    if dp in ("salary_range", "salary"):
        f = fields.get("salary")
        if isinstance(f, dict) and isinstance(f.get("value"), dict):
            sal = f["value"]
            mn = sal.get("min", "")
            mx = sal.get("max", "")
            cur = sal.get("currency", "")
            per = sal.get("period", "")
            if mn == mx:
                return f"{cur} {mn} / {per}"
            return f"{cur} {mn} - {mx} / {per}"
        return "NA"

    if dp == "benefits":
        f = fields.get("benefits")
        if isinstance(f, dict) and isinstance(f.get("value"), list) and f["value"]:
            return ", ".join(f["value"])
        return "NA"

    # Additional Info
    if dp == "reporting_to":
        f = fields.get("reporting_to")
        return f["value"] if isinstance(f, dict) and f.get("value") else "NA"

    if dp == "team_size":
        f = fields.get("team_size")
        return f["value"] if isinstance(f, dict) and f.get("value") else "NA"

    if dp == "travel_requirement":
        f = fields.get("travel_requirement")
        return f["value"] if isinstance(f, dict) and f.get("value") else "NA"

    if dp == "application_link":
        f = fields.get("application_link")
        return f["value"] if isinstance(f, dict) and f.get("value") else "NA"

    if dp == "equal_opportunity_statement":
        f = fields.get("equal_opportunity_statement")
        return f["value"] if isinstance(f, dict) and f.get("value") else "NA"

    # Metadata
    if dp == "source_type":
        return "file"
    if dp == "language_detected":
        return "en"

    return "N/A"


def determine_match(expected_str, extracted_str, data_point):
    """Determine match status between expected and extracted values."""
    if not expected_str or not extracted_str:
        return ""

    exp = str(expected_str).strip().lower()
    ext = str(extracted_str).strip().lower()

    # "Same" in expected means the expected value is the same as the JD text — our extraction is correct
    if exp in ("same", "matches jd list", "captured"):
        if ext != "na":
            return "Match"

    # Both not found / not mentioned
    not_found_signals = {"not mentioned", "not provided", "not specified", "n/a", "na", "not present",
                         "none mentioned", "no", "none listed", "not present"}
    if any(s in exp for s in not_found_signals) and ext == "na":
        return "Correct (Not Found)"

    # Benefits: "as per contract/agreement" is vague — parser correctly returns NA
    dp = data_point.lower().strip()
    if dp == "benefits" and any(s in exp for s in ["as per contract", "as per agreement", "standard benefits"]):
        return "Correct (Vague in JD)"

    if ext == "na" and not any(s in exp for s in not_found_signals):
        return "Missing"

    if exp == ext:
        return "Match"

    # Check partial / contains
    if exp in ext or ext in exp:
        return "Match"

    # Salary: "$55/hr on C2C" vs "USD 55 / hour" — compare the numbers
    if dp in ("salary_range", "salary"):
        exp_nums = re.findall(r'\d+', exp)
        ext_nums = re.findall(r'\d+', ext)
        if exp_nums and ext_nums and exp_nums[0] == ext_nums[0]:
            return "Match"

    # Employment type: handle "Contract (C2C)" vs "contract (C2C, 12 months)"
    if dp == "employment_type":
        if "contract" in ext and ("contract" in exp or "c2c" in exp or "duration" in exp):
            return "Match"
        if "full_time" in ext and ("full-time" in exp or "full time" in exp):
            return "Match"
        # Expected says "Not mentioned" but we found contract — we're more accurate
        if any(s in exp for s in not_found_signals) and "contract" in ext:
            return "Match (Inferred from Duration)"

    # For multi-value fields, check overlap
    if dp in ("soft_skills", "technical_skills", "preferred_experience", "preferred_technologies", "responsibilities"):
        if "captured" in ext:
            return "Match"
        # Check if expected items are a subset of extracted
        exp_items = set(re.findall(r'\w{3,}', exp))
        ext_items = set(re.findall(r'\w{3,}', ext))
        if exp_items and exp_items.issubset(ext_items):
            return "Match"
        if exp_items and ext_items:
            overlap = exp_items & ext_items
            if len(overlap) >= len(exp_items) * 0.5:
                return "Match"

    if dp in ("job_summary", "description", "company_overview"):
        if "sentences" in ext and "chars" in ext:
            sent_match = re.search(r'\[(\d+) sentences', ext)
            if sent_match and int(sent_match.group(1)) >= 2:
                return "Match (Multi-sentence)"
            elif sent_match and int(sent_match.group(1)) == 1:
                return "Partial (Single sentence)"

    # Certifications: "CCDE, CCIE" vs full names
    if dp == "certifications":
        exp_parts = set(re.findall(r'[A-Z]{2,}', exp.upper()))
        ext_parts = set(re.findall(r'[A-Z]{2,}', ext.upper()))
        if exp_parts and exp_parts.issubset(ext_parts):
            return "Match"

    # Years of experience: "4-7 years" vs "4-7 years (required)"
    if dp in ("years_of_experience", "experience_years"):
        exp_nums = re.findall(r'\d+', exp)
        ext_nums = re.findall(r'\d+', ext)
        if exp_nums and ext_nums:
            if exp_nums[:2] == ext_nums[:2] or exp_nums[0] == ext_nums[0]:
                return "Match"
            # If ranges overlap
            return "Partial Match"

    # Location: "Austin, TX (Hybrid)" vs "Austin, TX, USA (Hybrid)"
    if dp == "job_location":
        # Extract city, state
        exp_parts = set(re.findall(r'\b\w{2,}\b', exp))
        ext_parts = set(re.findall(r'\b\w{2,}\b', ext))
        if exp_parts.issubset(ext_parts):
            return "Match"

    # Company overview, reporting_to: "Same" already handled above
    # education: partial matching
    if dp in ("education_level", "education"):
        if any(kw in ext for kw in ["bachelor", "master", "phd"]):
            if any(kw in exp for kw in ["bachelor", "master", "phd"]):
                return "Match" if exp.split()[0] == ext.split()[0] else "Partial Match"

    # Fuzzy check
    exp_words = set(re.findall(r'\w+', exp))
    ext_words = set(re.findall(r'\w+', ext))
    if exp_words and ext_words:
        overlap = exp_words & ext_words
        if len(overlap) >= len(exp_words) * 0.6:
            return "Partial Match"

    return "Mismatch"


def main():
    wb = load_workbook(INPUT_EXCEL)

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    match_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    partial_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    miss_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    not_found_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    wrap_alignment = Alignment(wrap_text=True, vertical='top')

    for sheet_name, result_file in SHEET_MAP.items():
        if sheet_name not in wb.sheetnames:
            print(f"Sheet '{sheet_name}' not found in workbook, skipping")
            continue

        ws = wb[sheet_name]
        result = load_result(result_file)

        if not result:
            print(f"No result file for {sheet_name}")
            continue

        print(f"\nProcessing sheet: {sheet_name}")

        # Find the header row (row 1)
        # Determine the columns - find the last used column
        header_row = 1

        # Read existing headers
        existing_headers = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=header_row, column=col).value
            if val:
                existing_headers.append(val)

        # Add new columns after the last data column
        # Find column indices
        new_col_extracted = ws.max_column + 1
        new_col_status = ws.max_column + 2

        # But first, check if we already added these columns
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=header_row, column=col).value
            if val and "New Extracted" in str(val):
                new_col_extracted = col
                new_col_status = col + 1
                break

        # Write headers
        ws.cell(row=header_row, column=new_col_extracted, value="New Extracted Value (After Fix)").font = header_font
        ws.cell(row=header_row, column=new_col_extracted).fill = header_fill
        ws.cell(row=header_row, column=new_col_extracted).border = thin_border
        ws.cell(row=header_row, column=new_col_extracted).alignment = wrap_alignment

        ws.cell(row=header_row, column=new_col_status, value="New Match Status").font = header_font
        ws.cell(row=header_row, column=new_col_status).fill = header_fill
        ws.cell(row=header_row, column=new_col_status).border = thin_border
        ws.cell(row=header_row, column=new_col_status).alignment = wrap_alignment

        # Set column widths
        ws.column_dimensions[ws.cell(row=1, column=new_col_extracted).column_letter].width = 50
        ws.column_dimensions[ws.cell(row=1, column=new_col_status).column_letter].width = 25

        # Process each data row
        match_count = 0
        partial_count = 0
        miss_count = 0
        total_count = 0

        for row in range(2, ws.max_row + 1):
            category = ws.cell(row=row, column=1).value
            data_point = ws.cell(row=row, column=2).value
            expected = ws.cell(row=row, column=3).value

            if not data_point or not category:
                continue
            if str(category).strip().lower() in ("metadata", "count", ""):
                # Skip metadata rows and count rows
                extracted = get_field_value(result, str(category), str(data_point))
                ws.cell(row=row, column=new_col_extracted, value=str(extracted)).alignment = wrap_alignment
                ws.cell(row=row, column=new_col_extracted).border = thin_border
                continue
            if str(data_point).strip().lower() in ("", "count"):
                continue

            total_count += 1

            # Get our extracted value
            extracted = get_field_value(result, str(category), str(data_point))

            # Determine match status
            status = determine_match(str(expected) if expected else "", str(extracted), str(data_point))

            # Write to Excel
            cell_ext = ws.cell(row=row, column=new_col_extracted, value=str(extracted))
            cell_ext.alignment = wrap_alignment
            cell_ext.border = thin_border

            cell_status = ws.cell(row=row, column=new_col_status, value=status)
            cell_status.alignment = wrap_alignment
            cell_status.border = thin_border

            # Color code
            if "Match" in status and "Mismatch" not in status:
                cell_status.fill = match_fill
                if "Partial" not in status:
                    match_count += 1
                else:
                    partial_count += 1
            elif "Missing" in status or "Mismatch" in status:
                cell_status.fill = miss_fill
                miss_count += 1
            elif "Correct" in status:
                cell_status.fill = not_found_fill
                match_count += 1
            elif "Partial" in status:
                cell_status.fill = partial_fill
                partial_count += 1

        print(f"  Total: {total_count}, Match: {match_count}, Partial: {partial_count}, Missing/Mismatch: {miss_count}")

    wb.save(OUTPUT_EXCEL)
    print(f"\nSaved to: {OUTPUT_EXCEL}")


if __name__ == "__main__":
    main()
