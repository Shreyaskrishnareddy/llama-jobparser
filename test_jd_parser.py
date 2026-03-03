"""
JD Parser — 3-Cycle Comprehensive Test Suite
Cycle 1: Parse all 35 JDs, build ground truth, identify issues
Cycle 2: Retest after fixes
Cycle 3: Final regression + professional Excel report

Usage:
    python3 test_jd_parser.py --cycle 1   # Parse all, discover issues
    python3 test_jd_parser.py --cycle 2   # Retest after fixes
    python3 test_jd_parser.py --cycle 3   # Final regression + report
"""

import argparse
import glob
import json
import os
import re
import sys
import time
from datetime import datetime

from dotenv import load_dotenv
load_dotenv()

import groq_jd_parser
groq_jd_parser.GROQ_API_KEY = os.environ.get("GROQ_API_KEY", "")
groq_jd_parser.GROQ_MODEL = os.environ.get("GROQ_MODEL", "llama-3.1-8b-instant")

from groq_jd_parser import parse_jd, extract_text_from_file, _ALL_FIELDS

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

PDF_DIR = "/mnt/d/DOWNLOADS/JD Parser-20260302T072757Z-1-001/JD Parser/JD PDF/"
DOCX_DIR = "/mnt/d/DOWNLOADS/JD Parser-20260302T072757Z-1-001/JD Parser/JD Word Format/"
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ALL_FIELDS = list(_ALL_FIELDS)

# Developer mode: shorter delays (higher rate limits)
DELAY_BETWEEN_CALLS = 3
MAX_RETRIES_PER_JD = 3


# ---------------------------------------------------------------------------
# File discovery
# ---------------------------------------------------------------------------

def discover_jd_files():
    files = []
    for fp in sorted(glob.glob(os.path.join(PDF_DIR, "*.pdf"))):
        files.append({"filename": os.path.basename(fp), "filepath": fp, "format": "PDF"})
    for fp in sorted(glob.glob(os.path.join(DOCX_DIR, "*.docx"))):
        files.append({"filename": os.path.basename(fp), "filepath": fp, "format": "DOCX"})
    return files


# ---------------------------------------------------------------------------
# Result management
# ---------------------------------------------------------------------------

def _safe_name(filename):
    return filename.replace(" ", "_").replace("/", "_").replace("(", "").replace(")", "").replace("&", "and")


def get_results_dir(cycle):
    d = os.path.join(BASE_DIR, f"test-results/cycle{cycle}")
    os.makedirs(d, exist_ok=True)
    return d


def get_result_path(cycle, filename):
    return os.path.join(get_results_dir(cycle), f"{_safe_name(filename)}.json")


def is_already_processed(cycle, filename):
    path = get_result_path(cycle, filename)
    if not os.path.exists(path):
        return False
    try:
        with open(path) as f:
            data = json.load(f)
        return "fields" in data and "error" not in data
    except (json.JSONDecodeError, IOError):
        return False


def load_cached_result(cycle, filename):
    path = get_result_path(cycle, filename)
    if not os.path.exists(path):
        return None
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def save_result(cycle, filename, result):
    path = get_result_path(cycle, filename)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2, ensure_ascii=False, default=str)


# ---------------------------------------------------------------------------
# Process single JD (with retry)
# ---------------------------------------------------------------------------

def process_single_jd(jd_info):
    filename = jd_info["filename"]
    filepath = jd_info["filepath"]
    fmt = jd_info["format"]

    result = {
        "filename": filename,
        "format": fmt,
        "status": "pending",
        "raw_text": "",
        "parse_result": None,
        "fields_filled": 0,
        "total_fields": len(ALL_FIELDS),
        "global_confidence": 0.0,
        "processing_time_ms": 0,
        "error": None,
        "retries": 0,
        "field_details": {},
    }

    start = time.time()

    # Extract text
    try:
        raw_text = extract_text_from_file(filepath)
        result["raw_text"] = raw_text or ""
    except Exception as e:
        result["status"] = "fail"
        result["error"] = f"Text extraction failed: {e}"
        result["processing_time_ms"] = int((time.time() - start) * 1000)
        return result

    if not raw_text or len(raw_text.strip()) < 30:
        result["status"] = "fail"
        result["error"] = "Extracted text too short or empty"
        result["processing_time_ms"] = int((time.time() - start) * 1000)
        return result

    # Parse with retries
    parse_output = None
    last_error = None
    for attempt in range(MAX_RETRIES_PER_JD):
        try:
            parse_output = parse_jd(raw_text, filename=filename)
            if "error" not in parse_output:
                break  # Success
            last_error = parse_output.get("error", "Unknown")
            result["retries"] = attempt + 1
            if attempt < MAX_RETRIES_PER_JD - 1:
                time.sleep(2)  # Brief wait before retry
        except Exception as e:
            last_error = f"parse_jd exception: {e}"
            result["retries"] = attempt + 1
            if attempt < MAX_RETRIES_PER_JD - 1:
                time.sleep(2)

    result["processing_time_ms"] = int((time.time() - start) * 1000)

    if parse_output is None or "error" in parse_output:
        result["status"] = "fail"
        result["error"] = last_error
        if parse_output:
            result["parse_result"] = parse_output
        return result

    result["status"] = "success"
    result["parse_result"] = parse_output
    result["global_confidence"] = parse_output.get("global_confidence", 0.0)

    # Detailed field analysis
    fields = parse_output.get("fields", {})
    filled = 0
    for fname in ALL_FIELDS:
        fdata = fields.get(fname)
        if fdata is not None:
            filled += 1
            val = fdata.get("value")
            conf = fdata.get("confidence", 0.0)
            has_prov = bool(fdata.get("provenance", {}).get("spans"))
            result["field_details"][fname] = {
                "status": "filled",
                "confidence": conf,
                "has_provenance": has_prov,
                "value_type": type(val).__name__,
                "value_preview": _preview_value(val),
            }
        else:
            result["field_details"][fname] = {
                "status": "missing",
                "confidence": 0.0,
                "has_provenance": False,
                "value_type": "NoneType",
                "value_preview": "",
            }

    result["fields_filled"] = filled
    return result


def _preview_value(val):
    """Create a short preview of a field value."""
    if val is None:
        return ""
    if isinstance(val, str):
        return val[:100] + ("..." if len(val) > 100 else "")
    if isinstance(val, list):
        if len(val) == 0:
            return "[]"
        return f"[{len(val)} items]"
    if isinstance(val, dict):
        return json.dumps(val, ensure_ascii=False, default=str)[:100]
    return str(val)[:100]


# ---------------------------------------------------------------------------
# Test runner
# ---------------------------------------------------------------------------

def run_all_tests(jd_files, cycle, force=False):
    results = []
    total = len(jd_files)

    for i, jd_info in enumerate(jd_files, 1):
        filename = jd_info["filename"]

        if not force and is_already_processed(cycle, filename):
            print(f"[{i}/{total}] SKIP (cached): {filename}")
            cached = load_cached_result(cycle, filename)
            if cached:
                results.append({
                    "filename": cached.get("filename", filename),
                    "format": cached.get("format", jd_info["format"]),
                    "status": cached.get("status", "success"),
                    "raw_text": cached.get("raw_text", ""),
                    "parse_result": cached,
                    "fields_filled": cached.get("fields_filled", 0),
                    "total_fields": len(ALL_FIELDS),
                    "global_confidence": cached.get("global_confidence", 0.0),
                    "processing_time_ms": cached.get("processing_time_ms", 0),
                    "error": cached.get("error"),
                    "retries": cached.get("retries", 0),
                    "field_details": cached.get("field_details", {}),
                })
            continue

        print(f"[{i}/{total}] Processing: {filename}")
        result = process_single_jd(jd_info)

        # Save full result (parse_result merged in)
        save_data = dict(result)
        if result["parse_result"]:
            save_data.update(result["parse_result"])
        save_data.pop("parse_result", None)
        save_result(cycle, filename, save_data)
        results.append(result)

        icon = "OK" if result["status"] == "success" else "FAIL"
        retries_str = f" (retries: {result['retries']})" if result["retries"] > 0 else ""
        print(f"         -> {icon} | fields: {result['fields_filled']}/{len(ALL_FIELDS)} "
              f"| confidence: {result['global_confidence']:.2f} "
              f"| time: {result['processing_time_ms']}ms{retries_str}")
        if result["error"]:
            print(f"         -> ERROR: {result['error'][:200]}")

        if i < total:
            time.sleep(DELAY_BETWEEN_CALLS)

    return results


# ---------------------------------------------------------------------------
# Analysis: identify issues
# ---------------------------------------------------------------------------

def analyze_results(results):
    """Analyze results and print issue summary."""
    print("\n" + "=" * 70)
    print("ANALYSIS")
    print("=" * 70)

    total = len(results)
    success = [r for r in results if r["status"] == "success"]
    failed = [r for r in results if r["status"] != "success"]

    print(f"\nParse Rate: {len(success)}/{total} ({len(success)/total*100:.1f}%)")
    if failed:
        print(f"\nFailed JDs ({len(failed)}):")
        for r in failed:
            print(f"  - {r['filename']}: {r.get('error', 'unknown')[:100]}")

    if not success:
        return

    # Field coverage analysis
    print(f"\n--- Field Coverage (across {len(success)} successful parses) ---")
    field_stats = {}
    for fname in ALL_FIELDS:
        filled_count = sum(1 for r in success if r.get("field_details", {}).get(fname, {}).get("status") == "filled")
        field_stats[fname] = {
            "filled": filled_count,
            "missing": len(success) - filled_count,
            "pct": filled_count / len(success) * 100,
        }

    # Sort by fill rate
    always_filled = []
    sometimes_filled = []
    rarely_filled = []
    for fname, stats in sorted(field_stats.items(), key=lambda x: x[1]["pct"], reverse=True):
        if stats["pct"] == 100:
            always_filled.append(fname)
        elif stats["pct"] > 50:
            sometimes_filled.append((fname, stats["pct"]))
        else:
            rarely_filled.append((fname, stats["pct"]))

    print(f"\n  Always filled ({len(always_filled)}/{len(ALL_FIELDS)}): {', '.join(always_filled)}")
    if sometimes_filled:
        print(f"\n  Sometimes filled:")
        for fname, pct in sometimes_filled:
            print(f"    {fname}: {pct:.0f}%")
    if rarely_filled:
        print(f"\n  Rarely/Never filled:")
        for fname, pct in rarely_filled:
            print(f"    {fname}: {pct:.0f}%")

    # Confidence analysis
    all_confs = [r["global_confidence"] for r in success]
    avg_conf = sum(all_confs) / len(all_confs)
    min_conf = min(all_confs)
    max_conf = max(all_confs)
    print(f"\n--- Confidence ---")
    print(f"  Average: {avg_conf:.4f}")
    print(f"  Range: {min_conf:.4f} - {max_conf:.4f}")

    # Low confidence JDs
    low_conf = [r for r in success if r["global_confidence"] < 0.90]
    if low_conf:
        print(f"\n  Low confidence JDs (<0.90):")
        for r in low_conf:
            print(f"    {r['filename']}: {r['global_confidence']:.4f}")

    # PDF vs DOCX comparison
    pdf_results = [r for r in success if r["format"] == "PDF"]
    docx_results = [r for r in success if r["format"] == "DOCX"]
    if pdf_results and docx_results:
        print(f"\n--- Format Comparison ---")
        pdf_avg_fields = sum(r["fields_filled"] for r in pdf_results) / len(pdf_results)
        docx_avg_fields = sum(r["fields_filled"] for r in docx_results) / len(docx_results)
        pdf_avg_conf = sum(r["global_confidence"] for r in pdf_results) / len(pdf_results)
        docx_avg_conf = sum(r["global_confidence"] for r in docx_results) / len(docx_results)
        print(f"  PDF:  avg {pdf_avg_fields:.1f} fields | conf {pdf_avg_conf:.4f} | {len(pdf_results)} files")
        print(f"  DOCX: avg {docx_avg_fields:.1f} fields | conf {docx_avg_conf:.4f} | {len(docx_results)} files")

    # Cross-format consistency check
    print(f"\n--- Cross-Format Consistency ---")
    pdf_names = {r["filename"].replace(".pdf", "").lower(): r for r in success if r["format"] == "PDF"}
    docx_names = {r["filename"].replace(".docx", "").lower(): r for r in success if r["format"] == "DOCX"}
    common = set(pdf_names.keys()) & set(docx_names.keys())
    inconsistencies = []
    for name in sorted(common):
        pdf_r = pdf_names[name]
        docx_r = docx_names[name]
        pdf_fields = pdf_r.get("fields_filled", 0)
        docx_fields = docx_r.get("fields_filled", 0)
        diff = abs(pdf_fields - docx_fields)
        if diff > 2:
            inconsistencies.append((name, pdf_fields, docx_fields, diff))
    if inconsistencies:
        print(f"  Significant field differences (>2 fields):")
        for name, pf, df, diff in inconsistencies:
            print(f"    {name}: PDF={pf}, DOCX={df} (diff={diff})")
    else:
        print(f"  All {len(common)} common JDs are consistent between PDF and DOCX")

    # Processing time
    times = [r["processing_time_ms"] for r in success]
    avg_time = sum(times) / len(times)
    print(f"\n--- Performance ---")
    print(f"  Average: {avg_time:.0f}ms | Min: {min(times)}ms | Max: {max(times)}ms")

    return field_stats


# ---------------------------------------------------------------------------
# Excel report (professional)
# ---------------------------------------------------------------------------

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
SUCCESS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
MISSING_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)


def _style_header(ws, headers, widths):
    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"


def _add_autofilter(ws, num_cols):
    ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{ws.max_row}"


def _build_overview_sheet(wb, results, cycle):
    """Executive summary sheet."""
    ws = wb.active
    ws.title = "Executive Summary"

    success = [r for r in results if r["status"] == "success"]
    failed = [r for r in results if r["status"] != "success"]
    total = len(results)

    # Title
    ws.merge_cells("A1:F1")
    cell = ws.cell(row=1, column=1, value="JD Parser — Test Report")
    cell.font = Font(bold=True, size=16, color="003366")
    cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:F2")
    cell = ws.cell(row=2, column=1, value=f"Cycle {cycle} | {datetime.now().strftime('%B %d, %Y %H:%M')} | Model: {groq_jd_parser.GROQ_MODEL}")
    cell.font = Font(size=11, color="666666")
    cell.alignment = Alignment(horizontal="center")

    # Key metrics
    row = 4
    metrics = [
        ("Total JDs Tested", total),
        ("Parse Success Rate", f"{len(success)}/{total} ({len(success)/total*100:.1f}%)"),
        ("Parse Failures", f"{len(failed)}/{total}"),
        ("Average Fields Filled", f"{sum(r['fields_filled'] for r in success)/max(len(success),1):.1f} / {len(ALL_FIELDS)}"),
        ("Average Confidence Score", f"{sum(r['global_confidence'] for r in success)/max(len(success),1):.4f}"),
        ("Average Processing Time", f"{sum(r['processing_time_ms'] for r in success)/max(len(success),1):.0f}ms"),
    ]

    ws.cell(row=row, column=1, value="KEY METRICS").font = Font(bold=True, size=12, color="003366")
    row += 1
    for label, value in metrics:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=str(value))
        row += 1

    # Format comparison
    row += 1
    ws.cell(row=row, column=1, value="FORMAT COMPARISON").font = Font(bold=True, size=12, color="003366")
    row += 1
    for fmt_label, fmt_key in [("PDF", "PDF"), ("DOCX", "DOCX")]:
        fmt_results = [r for r in success if r["format"] == fmt_key]
        if fmt_results:
            avg_f = sum(r["fields_filled"] for r in fmt_results) / len(fmt_results)
            avg_c = sum(r["global_confidence"] for r in fmt_results) / len(fmt_results)
            ws.cell(row=row, column=1, value=fmt_label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=f"{len(fmt_results)} files | Avg {avg_f:.1f} fields | Conf {avg_c:.4f}")
            row += 1

    # Failures
    if failed:
        row += 1
        ws.cell(row=row, column=1, value="FAILURES").font = Font(bold=True, size=12, color="CC0000")
        row += 1
        for r in failed:
            ws.cell(row=row, column=1, value=r["filename"])
            ws.cell(row=row, column=2, value=r.get("error", "unknown")[:100])
            row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 20


def _build_summary_sheet(wb, results):
    ws = wb.create_sheet("JD Results")
    headers = ["#", "JD Filename", "Format", "Parse Status", "Fields Filled",
               "Fill Rate %", "Global Confidence", "Time (ms)", "Retries", "Error"]
    widths = [5, 55, 8, 14, 14, 12, 16, 12, 10, 60]
    _style_header(ws, headers, widths)

    for row_idx, r in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=r["filename"])
        ws.cell(row=row_idx, column=3, value=r["format"])
        status_cell = ws.cell(row=row_idx, column=4, value=r["status"])
        status_cell.fill = SUCCESS_FILL if r["status"] == "success" else FAIL_FILL
        ws.cell(row=row_idx, column=5, value=f"{r['fields_filled']} / {len(ALL_FIELDS)}")
        pct = round(r["fields_filled"] / len(ALL_FIELDS) * 100, 1) if r["fields_filled"] else 0
        ws.cell(row=row_idx, column=6, value=f"{pct}%")
        ws.cell(row=row_idx, column=7, value=round(r["global_confidence"], 4))
        ws.cell(row=row_idx, column=8, value=r["processing_time_ms"])
        ws.cell(row=row_idx, column=9, value=r.get("retries", 0))
        ws.cell(row=row_idx, column=10, value=r.get("error") or "")

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).border = THIN_BORDER

    # Summary row
    total_row = len(results) + 2
    success_count = sum(1 for r in results if r["status"] == "success")
    fail_count = len(results) - success_count
    ws.cell(row=total_row, column=1, value="").font = Font(bold=True)
    ws.cell(row=total_row, column=2, value="TOTALS").font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=f"{success_count} pass / {fail_count} fail").font = Font(bold=True)
    if success_count:
        avg_filled = sum(r["fields_filled"] for r in results if r["status"] == "success") / success_count
        avg_conf = sum(r["global_confidence"] for r in results if r["status"] == "success") / success_count
        avg_time = sum(r["processing_time_ms"] for r in results if r["status"] == "success") // success_count
        ws.cell(row=total_row, column=5, value=f"Avg: {avg_filled:.1f}").font = Font(bold=True)
        ws.cell(row=total_row, column=7, value=round(avg_conf, 4)).font = Font(bold=True)
        ws.cell(row=total_row, column=8, value=avg_time).font = Font(bold=True)

    _add_autofilter(ws, len(headers))


def _build_field_details_sheet(wb, results):
    ws = wb.create_sheet("Field Details")
    headers = ["JD Filename", "Format", "Field Name", "Extracted Value",
               "Confidence", "Status", "Has Provenance", "Value Type"]
    widths = [45, 8, 25, 80, 12, 14, 14, 12]
    _style_header(ws, headers, widths)

    row_idx = 2
    for r in results:
        filename = r["filename"]
        fmt = r["format"]

        if r["status"] != "success" or not r.get("parse_result"):
            ws.cell(row=row_idx, column=1, value=filename)
            ws.cell(row=row_idx, column=2, value=fmt)
            ws.cell(row=row_idx, column=3, value="ALL")
            ws.cell(row=row_idx, column=4, value=f"PARSE FAILED: {r.get('error', 'unknown')}")
            ws.cell(row=row_idx, column=5, value=0.0)
            cell = ws.cell(row=row_idx, column=6, value="fail")
            cell.fill = FAIL_FILL
            ws.cell(row=row_idx, column=7, value="No")
            row_idx += 1
            continue

        parse_data = r["parse_result"]
        fields = parse_data.get("fields", {})

        for field_name in ALL_FIELDS:
            field_data = fields.get(field_name)
            ws.cell(row=row_idx, column=1, value=filename)
            ws.cell(row=row_idx, column=2, value=fmt)
            ws.cell(row=row_idx, column=3, value=field_name)

            if field_data is None:
                ws.cell(row=row_idx, column=4, value="")
                ws.cell(row=row_idx, column=5, value=0.0)
                cell = ws.cell(row=row_idx, column=6, value="missing")
                cell.fill = MISSING_FILL
                ws.cell(row=row_idx, column=7, value="No")
                ws.cell(row=row_idx, column=8, value="null")
            else:
                value = field_data.get("value")
                value_str = json.dumps(value, ensure_ascii=False, default=str)
                if len(value_str) > 32000:
                    value_str = value_str[:32000] + "...[truncated]"
                ws.cell(row=row_idx, column=4, value=value_str)

                conf = field_data.get("confidence", 0.0)
                ws.cell(row=row_idx, column=5, value=round(conf, 2))

                status = field_data.get("status", "unknown")
                status_cell = ws.cell(row=row_idx, column=6, value=status)
                if status == "ok":
                    status_cell.fill = SUCCESS_FILL
                elif status == "low_confidence":
                    status_cell.fill = WARN_FILL

                has_prov = "Yes" if field_data.get("provenance", {}).get("spans") else "No"
                ws.cell(row=row_idx, column=7, value=has_prov)
                ws.cell(row=row_idx, column=8, value=type(value).__name__)

            row_idx += 1

    _add_autofilter(ws, len(headers))


def _build_field_coverage_sheet(wb, results):
    """Field-level coverage matrix: which fields are filled for which JDs."""
    ws = wb.create_sheet("Field Coverage Matrix")

    success = [r for r in results if r["status"] == "success"]
    if not success:
        ws.cell(row=1, column=1, value="No successful parses to analyze")
        return

    # Headers: Field name + each JD
    ws.cell(row=1, column=1, value="Field Name").font = HEADER_FONT
    ws.cell(row=1, column=1).fill = HEADER_FILL
    ws.cell(row=1, column=1).border = THIN_BORDER
    ws.column_dimensions["A"].width = 25

    # Last column: fill rate
    last_col = len(success) + 2
    ws.cell(row=1, column=last_col, value="Fill Rate").font = HEADER_FONT
    ws.cell(row=1, column=last_col).fill = HEADER_FILL
    ws.column_dimensions[get_column_letter(last_col)].width = 12

    for col_idx, r in enumerate(success, 2):
        short_name = r["filename"][:20]
        cell = ws.cell(row=1, column=col_idx, value=short_name)
        cell.font = Font(bold=True, size=8)
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", size=8)
        cell.alignment = Alignment(text_rotation=90, horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = 4

    for row_idx, fname in enumerate(ALL_FIELDS, 2):
        ws.cell(row=row_idx, column=1, value=fname).font = Font(size=9)
        filled_count = 0
        for col_idx, r in enumerate(success, 2):
            fd = r.get("field_details", {}).get(fname, {})
            is_filled = fd.get("status") == "filled"
            cell = ws.cell(row=row_idx, column=col_idx, value="Y" if is_filled else "")
            cell.fill = SUCCESS_FILL if is_filled else MISSING_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(size=8)
            if is_filled:
                filled_count += 1

        pct = filled_count / len(success) * 100
        ws.cell(row=row_idx, column=last_col, value=f"{pct:.0f}%")

    ws.freeze_panes = "B2"


def _build_raw_text_sheet(wb, results):
    ws = wb.create_sheet("Raw Text")
    headers = ["JD Filename", "Format", "Text Length", "Raw Extracted Text"]
    widths = [45, 8, 14, 120]
    _style_header(ws, headers, widths)

    for row_idx, r in enumerate(results, 2):
        raw = r.get("raw_text", "")
        display = raw[:32000] + ("...[truncated]" if len(raw) > 32000 else "")
        ws.cell(row=row_idx, column=1, value=r["filename"])
        ws.cell(row=row_idx, column=2, value=r["format"])
        ws.cell(row=row_idx, column=3, value=len(raw))
        cell = ws.cell(row=row_idx, column=4, value=display)
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def generate_excel_report(results, output_path, cycle):
    wb = openpyxl.Workbook()
    _build_overview_sheet(wb, results, cycle)
    _build_summary_sheet(wb, results)
    _build_field_details_sheet(wb, results)
    _build_field_coverage_sheet(wb, results)
    _build_raw_text_sheet(wb, results)
    wb.save(output_path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="JD Parser Test Suite")
    parser.add_argument("--cycle", type=int, choices=[1, 2, 3], required=True,
                        help="Test cycle: 1=discover, 2=retest, 3=final+report")
    parser.add_argument("--force", action="store_true",
                        help="Force re-parse even if cached results exist")
    args = parser.parse_args()

    cycle = args.cycle
    results_dir = get_results_dir(cycle)
    excel_path = os.path.join(BASE_DIR, f"test-results/JD_Parser_Test_Report_Cycle{cycle}.xlsx")

    print("=" * 70)
    print(f"JD PARSER — CYCLE {cycle} TEST")
    print(f"Timestamp: {datetime.now().isoformat()}")
    print(f"Model: {groq_jd_parser.GROQ_MODEL}")
    print(f"API Key: {'configured' if groq_jd_parser.GROQ_API_KEY else 'MISSING'}")
    print(f"Results: {results_dir}")
    print("=" * 70)

    jd_files = discover_jd_files()
    pdf_count = sum(1 for f in jd_files if f["format"] == "PDF")
    docx_count = sum(1 for f in jd_files if f["format"] == "DOCX")
    print(f"\nDiscovered {len(jd_files)} JD files: {pdf_count} PDFs, {docx_count} DOCXs")
    print(f"Delay between calls: {DELAY_BETWEEN_CALLS}s | Max retries per JD: {MAX_RETRIES_PER_JD}")
    print(f"\nStarting Cycle {cycle}...\n")

    results = run_all_tests(jd_files, cycle, force=args.force)

    success = sum(1 for r in results if r["status"] == "success")
    fail = len(results) - success
    print(f"\n{'=' * 70}")
    print(f"CYCLE {cycle} COMPLETE: {success}/{len(results)} passed, {fail} failed")
    if success > 0:
        avg_conf = sum(r["global_confidence"] for r in results if r["status"] == "success") / success
        avg_time = sum(r["processing_time_ms"] for r in results if r["status"] == "success") / success
        avg_fields = sum(r["fields_filled"] for r in results if r["status"] == "success") / success
        print(f"Average fields filled: {avg_fields:.1f} / {len(ALL_FIELDS)}")
        print(f"Average confidence: {avg_conf:.4f}")
        print(f"Average processing time: {avg_time:.0f}ms")

    # Analysis
    field_stats = analyze_results(results)

    # Excel report
    print(f"\nGenerating Excel report...")
    generate_excel_report(results, excel_path, cycle)
    print(f"Report saved: {excel_path}")
    print(f"JSON results: {results_dir}/")


if __name__ == "__main__":
    main()
